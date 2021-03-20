import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import Workbook
from pprint import pprint

wb = openpyxl.open("phones.xlsx", read_only=True)
sheet = wb.active
cells = sheet['B2':'K21']
nameList    = []
XiaomiList  = []
AppleList   = []
SamsungList = []
HuaweiList  = []
NokiaList   = []
OppoList    = []
OnePlusList = []
RealmeList  = []
GoogleList  = []
MeizuList   = []
for i in range(2,sheet.max_row + 1):
    XiaomiList.append(sheet[i][1].value)

for i in range (2, 24):
    for j in range (10):
        print(sheet[i][j].value)

"""for i in range(2,sheet.max_row + 1):
    AppleList.append(sheet[i][1].value)

for i in range(2,sheet.max_row + 1):
    SamsungList.append(sheet[i][2].value)

for i in range(2,sheet.max_row + 1):
    HuaweiList.append(sheet[i][3].value)

for i in range(2,sheet.max_row + 1):
    NokiaList.append(sheet[i][4].value)

for i in range(2,sheet.max_row + 1):
    OppoList.append(sheet[i][5].value)

for i in range(2,sheet.max_row + 1):
    OnePlusList.append(sheet[i][6].value)
    
for i in range(2,sheet.max_row + 1):
    RealmeList.append(sheet[i][7].value)

for i in range(2,sheet.max_row + 1):
    GoogleList.append(sheet[i][8].value)

for i in range(2,sheet.max_row + 1):
    MeizuList.append(sheet[i][9].value)

for i in range(2,sheet.max_row + 1):
    nameList.append(sheet[i][10].value)


print(nameList)"""
print(XiaomiList)





index = [0,1,2,3,4]  # по x
values = [5,7,3,4,6] # по y
plt.bar(index,values)
plt.show()
